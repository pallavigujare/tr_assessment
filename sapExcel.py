import pyautogui
import openpyxl

class SaplogonBOT(object):
    def __init__(self,usr,pw,inst):
        self.usrid=usr
        self.password=pw
        self.instance=inst
    
    def activate_preinvoke(self):
        pyautogui.hotkey('win','d')
    
    def activate_Saplogon(self):
        try:
            pyautogui.hotkey('win','r')
            pyautogui.typewrite('Saplogon')
            pyautogui.press('enter')
            return True
        except:
            return False
    
    def activate_postinvoke(self):
            pyautogui.hotkey('win','up')
   
    def open_instance(self):
        try:
            pyautogui.hotkey('ctrl','f')
            pyautogui.typewrite(self.instance)
            pyautogui.press('enter')
            return True
        except:
            return False
        
    def sap_login(self):
         pyautogui.typewrite(self.usrid)
         pyautogui.press('tab')
         pyautogui.typewrite(self.password)
         pyautogui.press('enter')
         
    def sap_fs10n(self):
         time.sleep(2)
         pyautogui.typewrite('fs10n')
         pyautogui.press('enter')
         
    def sap_accept(self):
         time.sleep(3)
         pyautogui.press('enter')
         time.sleep(3)
         pyautogui.hotkey('shift','F7')
         
    
    def sap_excel(self):
        path = "C:\\Users\\gujare\\Desktop\\demo.xlsx"
        # workbook object is created 
        wb_obj = openpyxl.load_workbook(path) 
  
        sheet_obj = wb_obj.active 
        max_row = sheet_obj.max_column 
  
        # Loop will print all columns name 
        for i in range(1, max_row + 1): 
             cell_obj = sheet_obj.cell(row = 2, column = i) 
             print(cell_obj.value,end = " ")
             cells = sheet_obj['A'+str(i):'E'+str(i)]
             
             time.sleep(2)
             pyautogui.typewrite(cells[0].value)#'2390999999') # cells[0].value
             pyautogui.hotkey('tab')
             pyautogui.hotkey('tab')
             time.sleep(5)
             pyautogui.typewrite(cells[1].value)#'AR00') #cells[1].value
             pyautogui.hotkey('down')
             time.sleep(3)
             pyautogui.typewrite(cells[2].value)#'2019')
             pyautogui.hotkey('down')
             time.sleep(3)
             pyautogui.hotkey('down')
             time.sleep(3)
             pyautogui.typewrite(cells[3].value)#'10')
             time.sleep(3)
             pyautogui.hotkey('F8')
             time.sleep(3)

    def sap_screenshot(self):
         myScreenshot = pyautogui.screenshot()
         time.sleep(2)
         myScreenshot.save(r'C:\Users\gujare\Desktop\Check\AR11.jpg')
         time.sleep(5)
         pyautogui.hotkey('esc')
        
# ---------------------------------------------------------------------- #
        
import time
 
if __name__=='__main__':
     mysap = SaplogonBOT('usr','pw','inst')
     time.sleep(3)
     # mysap.SaplogonBOT('usr','pw','inst')
     mysap.activate_preinvoke()
     time.sleep(2)
     mysap.activate_Saplogon()
     time.sleep(7)
     mysap.activate_postinvoke()
     time.sleep(2)
     mysap.open_instance()
     time.sleep(5)
     mysap.sap_login()
     time.sleep(3)
     mysap.sap_fs10n()
     time.sleep(3)
     mysap.sap_accept()
     time.sleep(3)
     mysap.sap_excel()
     time.sleep(3)
     mysap.sap_screenshot()
     
     
     
     
     
        
         
         

